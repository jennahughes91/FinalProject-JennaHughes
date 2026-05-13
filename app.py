"""
Sprint Backlog Prioritization System
=====================================
Streamlit application — two-phase architecture:

  Phase 1 (runs once per file)   → Claude evaluates description quality
                                    and drafts improvements for weak items.
  Phase 2 (runs on every change) → Formula re-scores and re-ranks the
                                    backlog instantly whenever the user
                                    adjusts a business area or product
                                    team priority weight.

Setup
-----
1. Install dependencies:
       pip install -r requirements.txt

2. Set your Anthropic API key:
       export ANTHROPIC_API_KEY="sk-ant-..."
   OR create .streamlit/secrets.toml with:
       ANTHROPIC_API_KEY = "sk-ant-..."

3. Run:
       streamlit run app.py
"""

import hashlib
import json
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path

import anthropic
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

PROFILES_FILE = "priority_profiles.json"

# Claude models
MODEL_EVAL    = "claude-haiku-4-5-20251001"   # fast + cheap for batched evaluation
MODEL_SUGGEST = "claude-sonnet-4-6"           # better writing for suggestions

# How many items per Claude call
BATCH_SIZE = 25

# Formula weights (must sum to 1.0)
FORMULA_WEIGHTS = {
    "business_area": 0.45,
    "product_team":  0.35,
    "effort":        0.20,
}

# Flexible column name recognition — maps canonical name → accepted aliases
COLUMN_ALIASES = {
    "item_id":       ["item id", "id", "ticket", "story id", "issue id", "key", "issue key"],
    "title":         ["title", "summary", "name", "story", "story title"],
    "description":   ["description", "details", "body", "acceptance criteria", "user story", "story description"],
    "business_area": ["business area", "business unit", "domain", "area", "department", "ba"],
    "ba_priority":   ["business area priority", "ba priority", "business priority", "biz priority", "area priority"],
    "product_team":  ["product team", "team", "squad", "crew", "pod", "tribe"],
    "pt_priority":   ["product team priority", "team priority", "pt priority", "squad priority"],
    "effort":        ["effort", "story points", "sp", "points", "size", "estimate", "t-shirt size", "t-shirt", "tshirt"],
}

# T-shirt size → numeric effort (1=smallest, 5=largest)
TSHIRT_MAP = {
    "xs": 1, "xsmall": 1, "x-small": 1,
    "s":  2, "small":  2,
    "m":  3, "medium": 3, "med": 3,
    "l":  4, "large":  4,
    "xl": 5, "xlarge": 5, "x-large": 5,
}

# Priority label → numeric (1–5)
PRIORITY_LABEL_MAP = {
    "lowest": 1, "low": 1, "minor": 1,
    "medium": 2, "med": 2, "moderate": 2, "normal": 2,
    "high": 3, "major": 3,
    "highest": 4, "critical": 4, "urgent": 4,
    "blocker": 5, "showstopper": 5,
}


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSING HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def find_column(df: pd.DataFrame, canonical: str):
    """Return the actual column name in df that matches one of the canonical aliases, or None."""
    aliases = COLUMN_ALIASES.get(canonical, [canonical])
    lower_map = {c.strip().lower(): c for c in df.columns}
    for alias in aliases:
        if alias.lower() in lower_map:
            return lower_map[alias.lower()]
    return None


def normalize_priority(val, default: int = 3) -> int:
    """Convert a priority value (label or number) to an integer 1–5."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    s = str(val).strip().lower()
    if s in PRIORITY_LABEL_MAP:
        return PRIORITY_LABEL_MAP[s]
    # "P1"–"P5" style (P1 = highest)
    if s.startswith("p") and s[1:].isdigit():
        p = int(s[1:])
        return max(1, min(5, 6 - p))
    try:
        n = round(float(s))
        return max(1, min(5, n))
    except ValueError:
        return default


def normalize_effort(val) -> int:
    """Convert effort (T-shirt size or numeric story points) to 1–5 scale."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 3
    s = str(val).strip().lower()
    if s in TSHIRT_MAP:
        return TSHIRT_MAP[s]
    try:
        n = float(s)
        if n <= 1:   return 1
        elif n <= 3: return 2
        elif n <= 5: return 3
        elif n <= 8: return 4
        else:        return 5
    except ValueError:
        return 3


def parse_backlog(file_bytes: bytes):
    """
    Parse an uploaded backlog Excel file.
    Returns (DataFrame of normalized items, list of warning strings).
    """
    try:
        df_raw = pd.read_excel(BytesIO(file_bytes))
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    if df_raw.empty:
        raise ValueError("The uploaded file is empty.")

    warnings = []
    items = []

    for i, row in df_raw.iterrows():
        item = {}

        for field in COLUMN_ALIASES:
            col = find_column(df_raw, field)
            raw_val = row[col] if col else None
            item[field] = raw_val if not (isinstance(raw_val, float) and pd.isna(raw_val)) else None

        # Item ID
        if item.get("item_id") is None:
            item["item_id"] = f"ITEM-{i + 1}"
            warnings.append(f"Row {i + 1}: No Item ID column found — assigned '{item['item_id']}'.")
        else:
            item["item_id"] = str(item["item_id"]).strip()

        # Title
        if not item.get("title"):
            item["title"] = f"Untitled Item {i + 1}"
            warnings.append(f"Row {i + 1} ({item['item_id']}): Missing title.")
        else:
            item["title"] = str(item["title"]).strip()

        # Description
        item["description"] = str(item.get("description") or "").strip()
        if not item["description"]:
            warnings.append(f"Row {i + 1} ({item['item_id']}): Empty description — will be flagged as weak.")

        # Business Area
        item["business_area"] = str(item.get("business_area") or "Unknown").strip()
        item["ba_priority"] = normalize_priority(item.get("ba_priority"), default=3)

        # Product Team
        item["product_team"] = str(item.get("product_team") or "Unknown").strip()
        item["pt_priority"] = normalize_priority(item.get("pt_priority"), default=3)

        # Effort
        item["effort_raw"]  = item.get("effort")
        item["effort_norm"] = normalize_effort(item.get("effort"))

        items.append(item)

    return pd.DataFrame(items), warnings


# ─────────────────────────────────────────────────────────────────────────────
# RANKING FORMULA  (Phase 2 — pure math, runs on every slider change)
# ─────────────────────────────────────────────────────────────────────────────

def compute_scores(df: pd.DataFrame, ba_weights: dict, pt_weights: dict) -> pd.DataFrame:
    """
    Compute a composite priority score for each item and return a sorted DataFrame.

    Scoring:
        ba_score  = (user_ba_weight / 5) × (item_ba_priority / 5)   [0–1]
        pt_score  = (user_pt_weight / 5) × (item_pt_priority / 5)   [0–1]
        eff_score = (5 – effort_norm) / 4                            [0–1, inverse]

        priority_score = 0.45 × ba_score + 0.35 × pt_score + 0.20 × eff_score
    """
    w = FORMULA_WEIGHTS

    def _score(row):
        ba_user = ba_weights.get(row["business_area"], 3) / 5.0
        pt_user = pt_weights.get(row["product_team"],  3) / 5.0
        ba_s  = ba_user * (row["ba_priority"]  / 5.0)
        pt_s  = pt_user * (row["pt_priority"]  / 5.0)
        eff_s = (5 - row["effort_norm"])        / 4.0
        return round(
            w["business_area"] * ba_s +
            w["product_team"]  * pt_s +
            w["effort"]        * eff_s,
            4
        )

    result = df.copy()
    result["priority_score"] = result.apply(_score, axis=1)
    result = result.sort_values("priority_score", ascending=False).reset_index(drop=True)
    result["rank"] = result.index + 1
    return result


# ─────────────────────────────────────────────────────────────────────────────
# CLAUDE INTEGRATION  (Phase 1 — runs once per unique file)
# ─────────────────────────────────────────────────────────────────────────────

EVAL_SYSTEM_PROMPT = """You are a senior product manager evaluating development backlog item descriptions.

Evaluate each item against these four criteria:
1. BUSINESS PURPOSE — Does it explain why the work matters or what problem it solves?
2. SCOPE — Is it clear what is included and what is not?
3. OUTCOME — Are success criteria or expected outcomes stated or clearly implied?
4. ACTIONABILITY — Could a developer start work without asking clarifying questions?

A description is WEAK if it fails 2 or more criteria, or critically fails criterion 4 alone.
A description is STRONG if it meets at least 3 criteria.

Return ONLY a valid JSON array. Each element must have exactly these keys:
  - "item_id": string
  - "quality": "strong" | "weak"
  - "confidence": float 0.0–1.0
  - "reason": one sentence explaining the rating (for weak items, name the missing element)
  - "suggestion": for weak items, a rewritten description (2–4 sentences); empty string for strong items

--- WEAK EXAMPLES ---
  "Fix the login bug"                      → fails all criteria; no context, no scope, not actionable
  "Improve dashboard performance"          → vague; no target metric, no scope
  "Update the API as discussed in meeting" → relies on external context; not self-contained

--- STRONG EXAMPLES ---
  "Add email-based password reset so users who forget credentials can recover access without
   contacting support. Scope: email flow only, not SMS. Done when a user completes reset
   in under 2 minutes."
  "Migrate customer data exports to async job queue to prevent API timeouts on exports over
   10 000 rows. Success: zero HTTP timeouts for any export request."
"""


def _evaluate_batch(client: anthropic.Anthropic, batch: list) -> list:
    """Send one batch of items to Claude for description evaluation."""
    items_json = json.dumps(
        [{"item_id": it["item_id"], "title": it["title"], "description": it["description"]}
         for it in batch],
        indent=2
    )
    response = client.messages.create(
        model=MODEL_EVAL,
        max_tokens=4096,
        system=EVAL_SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": (
                f"Evaluate these {len(batch)} backlog items and return a JSON array.\n\n"
                f"{items_json}"
            )
        }]
    )
    text = response.content[0].text.strip()
    start = text.find("[")
    end   = text.rfind("]") + 1
    if start == -1 or end == 0:
        raise ValueError("No JSON array found in Claude response.")
    return json.loads(text[start:end])


def run_phase1_analysis(items_df: pd.DataFrame, progress_callback=None) -> dict:
    """
    Run the full Phase 1 AI pipeline.
    Returns a dict keyed by item_id, each value containing:
      quality, confidence, reason, suggestion
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY") or st.secrets.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        st.error(
            "**ANTHROPIC_API_KEY not found.** "
            "Set it as an environment variable or add it to `.streamlit/secrets.toml`."
        )
        st.stop()

    client  = anthropic.Anthropic(api_key=api_key)
    items   = items_df.to_dict("records")
    batches = [items[i:i + BATCH_SIZE] for i in range(0, len(items), BATCH_SIZE)]
    evaluations = {}

    for idx, batch in enumerate(batches):
        if progress_callback:
            progress_callback(
                f"Evaluating descriptions — batch {idx + 1} of {len(batches)}…",
                idx / len(batches)
            )
        try:
            results = _evaluate_batch(client, batch)
            for r in results:
                evaluations[str(r["item_id"])] = r
        except Exception as exc:
            for it in batch:
                evaluations[str(it["item_id"])] = {
                    "item_id":    it["item_id"],
                    "quality":    "review_needed",
                    "confidence": 0.0,
                    "reason":     f"Evaluation unavailable: {str(exc)[:120]}",
                    "suggestion": "",
                }

    if progress_callback:
        progress_callback("Analysis complete!", 1.0)

    return evaluations


# ─────────────────────────────────────────────────────────────────────────────
# PRIORITY PROFILE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def load_profiles() -> list:
    if not Path(PROFILES_FILE).exists():
        return []
    try:
        with open(PROFILES_FILE) as f:
            return json.load(f)
    except Exception:
        return []


def _save_profiles(profiles: list):
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2)


def create_profile(name: str, ba_weights: dict, pt_weights: dict, notes: str = "") -> dict:
    profiles = load_profiles()
    profile = {
        "id":         hashlib.md5(f"{name}{datetime.now().isoformat()}".encode()).hexdigest()[:10],
        "name":       name,
        "created_at": datetime.now().isoformat(),
        "ba_weights": dict(ba_weights),
        "pt_weights": dict(pt_weights),
        "notes":      notes,
    }
    profiles.append(profile)
    _save_profiles(profiles)
    return profile


def delete_profile(profile_id: str):
    profiles = [p for p in load_profiles() if p["id"] != profile_id]
    _save_profiles(profiles)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_export(ranked_df: pd.DataFrame, evaluations: dict) -> BytesIO:
    """Build and return a formatted Excel workbook as a BytesIO buffer."""
    output = BytesIO()
    rows = []
    for _, row in ranked_df.iterrows():
        iid = str(row["item_id"])
        ev  = evaluations.get(iid, {})
        rows.append({
            "Rank":                  int(row["rank"]),
            "Priority Score":        round(float(row["priority_score"]), 3),
            "Item ID":               iid,
            "Title":                 row["title"],
            "Business Area":         row["business_area"],
            "BA Item Priority":      int(row["ba_priority"]),
            "Product Team":          row["product_team"],
            "PT Item Priority":      int(row["pt_priority"]),
            "Effort":                row.get("effort_raw") or row["effort_norm"],
            "Description":           row["description"],
            "Description Quality":   (ev.get("quality") or "").replace("_", " ").title(),
            "Quality Confidence":    round(ev.get("confidence") or 0.0, 2),
            "Quality Reason":        ev.get("reason") or "",
            "Suggested Description": ev.get("suggestion") or "",
        })

    export_df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Prioritized Backlog")
        ws = writer.sheets["Prioritized Backlog"]

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        weak_fill   = PatternFill("solid", fgColor="FFF2CC")
        strong_fill = PatternFill("solid", fgColor="E2EFDA")
        review_fill = PatternFill("solid", fgColor="F2F2F2")

        for r_idx in range(2, ws.max_row + 1):
            quality_val = (ws.cell(r_idx, 11).value or "").lower()
            fill = (
                weak_fill   if quality_val == "weak"          else
                strong_fill if quality_val == "strong"        else
                review_fill if quality_val == "review needed" else None
            )
            if fill:
                ws.cell(r_idx, 11).fill = fill

        col_widths = [7, 13, 12, 36, 18, 14, 18, 14, 8, 55, 18, 16, 50, 65]
        for idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = w

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY
# ─────────────────────────────────────────────────────────────────────────────

def md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT APPLICATION
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Backlog Prioritization",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .badge-strong  { background:#E2EFDA; color:#276221; padding:2px 9px;
                     border-radius:4px; font-size:.83em; font-weight:600; }
    .badge-weak    { background:#FFF2CC; color:#7D5B00; padding:2px 9px;
                     border-radius:4px; font-size:.83em; font-weight:600; }
    .badge-review  { background:#F2F2F2; color:#555;    padding:2px 9px;
                     border-radius:4px; font-size:.83em; }
    div[data-testid="stMetric"] label { font-size:.8em !important; }
    .stExpander { border:1px solid #BDD7EE !important; border-radius:4px !important; }
</style>
""", unsafe_allow_html=True)

# ── Session state defaults ────────────────────────────────────────────────────
_DEFAULTS = {
    "items_df":       None,
    "evaluations":    {},
    "analysis_done":  False,
    "file_hash":      None,
    "ba_weights":     {},
    "pt_weights":     {},
    "parse_warnings": [],
    "active_profile": None,
    "weights_dirty":  False,
}
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📋 Backlog Prioritizer")
    st.caption("Powered by Anthropic Claude")
    st.divider()

    # ── 1. Upload ─────────────────────────────────────────────────────────────
    st.markdown("### 1 · Upload Backlog")
    backlog_file = st.file_uploader("Backlog Excel", type=["xlsx", "xls"])

    if backlog_file:
        raw = backlog_file.read()
        fh  = md5(raw)

        if fh != st.session_state.file_hash:
            with st.spinner("Parsing backlog…"):
                try:
                    df, warns = parse_backlog(raw)
                except ValueError as e:
                    st.error(str(e))
                    st.stop()

            st.session_state.items_df       = df
            st.session_state.file_hash      = fh
            st.session_state.parse_warnings = warns
            st.session_state.evaluations    = {}
            st.session_state.analysis_done  = False
            st.session_state.active_profile = None
            st.session_state.weights_dirty  = False
            st.session_state.ba_weights     = {a: 3 for a in df["business_area"].unique()}
            st.session_state.pt_weights     = {t: 3 for t in df["product_team"].unique()}
            st.rerun()

    # ── 2. AI Analysis ────────────────────────────────────────────────────────
    if st.session_state.items_df is not None and not st.session_state.analysis_done:
        st.divider()
        st.markdown("### 2 · AI Analysis")
        st.caption(
            "Claude evaluates every description and drafts improvements "
            "for weak items. Runs once — results are cached in this session."
        )
        n_items  = len(st.session_state.items_df)
        est_secs = max(10, n_items // 25 * 8)
        st.caption(f"~{est_secs}–{est_secs + 15} seconds for {n_items} items.")

        if st.button("▶  Analyze Descriptions", type="primary", use_container_width=True):
            prog_text = st.empty()
            prog_bar  = st.progress(0.0)

            evals = run_phase1_analysis(
                st.session_state.items_df,
                progress_callback=lambda msg, pct: (
                    prog_text.caption(msg),
                    prog_bar.progress(min(pct, 1.0))
                )
            )
            st.session_state.evaluations   = evals
            st.session_state.analysis_done = True
            prog_text.empty()
            prog_bar.empty()
            st.rerun()

    st.divider()

    # ── 3. Priority Weights ───────────────────────────────────────────────────
    if st.session_state.items_df is not None:
        st.markdown("### 3 · Priority Weights")
        st.caption("Adjust to re-rank instantly. Formula: 45% Business Area · 35% Product Team · 20% Effort")

        if st.session_state.active_profile:
            dirty = st.session_state.weights_dirty
            icon  = ":orange[●]" if dirty else "✅"
            note  = " *(unsaved changes)*" if dirty else ""
            st.markdown(f"**Profile:** {icon} *{st.session_state.active_profile}*{note}")
        else:
            st.markdown("**Profile:** —")

        with st.expander("🏢 Business Area Weights", expanded=True):
            new_ba = {}
            for area in sorted(st.session_state.ba_weights):
                new_ba[area] = st.slider(
                    area, min_value=1, max_value=5,
                    value=int(st.session_state.ba_weights.get(area, 3)),
                    key=f"ba__{area}"
                )
            if new_ba != st.session_state.ba_weights:
                st.session_state.ba_weights    = new_ba
                st.session_state.weights_dirty = True

        with st.expander("👥 Product Team Weights", expanded=True):
            new_pt = {}
            for team in sorted(st.session_state.pt_weights):
                new_pt[team] = st.slider(
                    team, min_value=1, max_value=5,
                    value=int(st.session_state.pt_weights.get(team, 3)),
                    key=f"pt__{team}"
                )
            if new_pt != st.session_state.pt_weights:
                st.session_state.pt_weights    = new_pt
                st.session_state.weights_dirty = True

        if st.button("↺  Reset all weights to 3", use_container_width=True):
            st.session_state.ba_weights    = {k: 3 for k in st.session_state.ba_weights}
            st.session_state.pt_weights    = {k: 3 for k in st.session_state.pt_weights}
            st.session_state.weights_dirty = True
            st.rerun()

        st.divider()

        # ── 4. Profiles ───────────────────────────────────────────────────────
        st.markdown("### 4 · Priority Profiles")
        profiles = load_profiles()

        if profiles:
            profile_map = {p["name"]: p for p in profiles}
            chosen = st.selectbox("Load a saved profile", ["— select —"] + list(profile_map))
            load_col, del_col = st.columns(2)

            with load_col:
                if st.button("Load", use_container_width=True):
                    if chosen != "— select —":
                        p = profile_map[chosen]
                        st.session_state.ba_weights     = dict(p["ba_weights"])
                        st.session_state.pt_weights     = dict(p["pt_weights"])
                        st.session_state.active_profile = p["name"]
                        st.session_state.weights_dirty  = False
                        st.rerun()

            with del_col:
                if st.button("Delete", use_container_width=True):
                    if chosen != "— select —":
                        delete_profile(profile_map[chosen]["id"])
                        if st.session_state.active_profile == chosen:
                            st.session_state.active_profile = None
                        st.rerun()

        with st.expander("💾 Save current weights as profile"):
            pname = st.text_input("Profile name", placeholder="e.g. Q2 Planning — Finance-led")
            notes = st.text_input("Notes (optional)")
            if st.button("Save Profile", type="primary", use_container_width=True):
                if pname.strip():
                    create_profile(pname.strip(), st.session_state.ba_weights,
                                   st.session_state.pt_weights, notes)
                    st.session_state.active_profile = pname.strip()
                    st.session_state.weights_dirty  = False
                    st.success(f"Saved profile: **{pname.strip()}**")
                else:
                    st.warning("Enter a profile name first.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CONTENT
# ══════════════════════════════════════════════════════════════════════════════

if st.session_state.items_df is None:
    st.markdown("# Sprint Backlog Prioritization")
    st.markdown(
        "Upload your **Backlog Excel** file in the sidebar to get started. "
        "The app will parse it automatically, let you adjust priority weights "
        "with live re-ranking, and optionally run AI description analysis via Claude."
    )
    st.info(
        "**Expected backlog columns** (flexible naming — common aliases are recognized automatically):\n\n"
        "| Field | Accepted column names |\n"
        "|-------|-----------------------|\n"
        "| Item ID | `Item ID`, `ID`, `Key`, `Ticket`, `Story ID` |\n"
        "| Title | `Title`, `Summary`, `Name`, `Story` |\n"
        "| Description | `Description`, `Details`, `Body`, `User Story` |\n"
        "| Business Area | `Business Area`, `Domain`, `Department`, `BA` |\n"
        "| Business Area Priority | `Business Area Priority`, `BA Priority` |\n"
        "| Product Team | `Product Team`, `Team`, `Squad`, `Pod` |\n"
        "| Product Team Priority | `Product Team Priority`, `Team Priority` |\n"
        "| Effort | `Effort`, `Story Points`, `SP`, `Size` (numeric or XS/S/M/L/XL) |\n"
    )
    st.stop()


# ── Live ranking (Phase 2) ────────────────────────────────────────────────────
ranked_df = compute_scores(
    st.session_state.items_df,
    st.session_state.ba_weights,
    st.session_state.pt_weights,
)

evals    = st.session_state.evaluations
n_total  = len(ranked_df)
n_evald  = len(evals)
n_weak   = sum(1 for e in evals.values() if e.get("quality") == "weak")
n_strong = sum(1 for e in evals.values() if e.get("quality") == "strong")

# ── Parse warnings ────────────────────────────────────────────────────────────
if st.session_state.parse_warnings:
    with st.expander(f"⚠️ {len(st.session_state.parse_warnings)} parse warning(s)", expanded=False):
        for w in st.session_state.parse_warnings:
            st.caption(f"• {w}")

# ── Summary metrics ───────────────────────────────────────────────────────────
m1, m2, m3, m4 = st.columns(4)
m1.metric("Total Items",            n_total)
m2.metric("Descriptions Analyzed",  n_evald)
m3.metric("Weak Descriptions",      n_weak,
          delta=f"{round(n_weak / n_evald * 100) if n_evald else 0}% of analyzed",
          delta_color="inverse")
m4.metric("Strong Descriptions",    n_strong)

# ── Export bar ────────────────────────────────────────────────────────────────
st.divider()
exp_col, lbl_col = st.columns([2, 6])
with exp_col:
    excel_buf = generate_export(ranked_df, evals)
    fname = f"prioritized_backlog_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button(
        "⬇  Export to Excel", data=excel_buf, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True
    )
with lbl_col:
    if not st.session_state.analysis_done:
        st.caption("💡 Run **AI Analysis** in the sidebar to add description quality flags to the export.")
    elif st.session_state.active_profile:
        st.caption(f"📌 Exporting with profile: **{st.session_state.active_profile}**")

st.divider()

# ── Filters ───────────────────────────────────────────────────────────────────
st.markdown(f"### Ranked Backlog  ·  {n_total} items")
fc1, fc2, fc3 = st.columns(3)
with fc1:
    area_filter = st.multiselect(
        "Filter: Business Area",
        sorted(ranked_df["business_area"].unique()), default=[]
    )
with fc2:
    team_filter = st.multiselect(
        "Filter: Product Team",
        sorted(ranked_df["product_team"].unique()), default=[]
    )
with fc3:
    qual_filter = st.multiselect(
        "Filter: Description Quality",
        ["strong", "weak", "review_needed"], default=[]
    )

# Apply filters
view_df = ranked_df.copy()
if area_filter:
    view_df = view_df[view_df["business_area"].isin(area_filter)]
if team_filter:
    view_df = view_df[view_df["product_team"].isin(team_filter)]
if qual_filter and evals:
    qual_lookup = {iid: e.get("quality", "") for iid, e in evals.items()}
    view_df = view_df[view_df["item_id"].astype(str).map(
        lambda iid: qual_lookup.get(iid, "") in qual_filter
    )]

if view_df.empty:
    st.info("No items match the current filters.")
    st.stop()

# ── Ranked item cards ─────────────────────────────────────────────────────────
for _, row in view_df.iterrows():
    iid     = str(row["item_id"])
    ev      = evals.get(iid, {})
    quality = ev.get("quality", "")
    reason  = ev.get("reason", "")
    suggest = ev.get("suggestion", "")
    ba_w    = st.session_state.ba_weights.get(row["business_area"], 3)
    pt_w    = st.session_state.pt_weights.get(row["product_team"],  3)

    badge = (
        '<span class="badge-strong">✓ Strong</span>' if quality == "strong"        else
        '<span class="badge-weak">⚠ Weak</span>'     if quality == "weak"          else
        '<span class="badge-review">? Review</span>' if quality == "review_needed" else ""
    )

    label = (
        f"#{int(row['rank'])}  {row['title']}  "
        f"·  {row['business_area']} (w{ba_w})  /  {row['product_team']} (w{pt_w})  "
        f"·  Score: {row['priority_score']:.3f}"
    )

    with st.expander(label, expanded=False):
        r1c1, r1c2, r1c3, r1c4 = st.columns([2, 3, 1, 3])
        r1c1.markdown(f"**ID:** `{iid}`")
        r1c2.markdown(
            f"**Business Area:** {row['business_area']}  "
            f"*(item priority {row['ba_priority']}, area weight {ba_w})*"
        )
        r1c3.markdown(f"**Effort:** {row.get('effort_raw') or row['effort_norm']}")
        r1c4.markdown(
            f"**Product Team:** {row['product_team']}  "
            f"*(item priority {row['pt_priority']}, team weight {pt_w})*"
        )

        st.markdown("**Description:**")
        st.text_area(
            "desc", value=row["description"], height=80,
            key=f"desc_{iid}", disabled=True, label_visibility="collapsed"
        )

        if quality:
            st.markdown(f"**Description Quality:** {badge}  {reason}", unsafe_allow_html=True)

        if suggest:
            st.markdown("**💡 Suggested Improvement:**")
            st.info(suggest)
